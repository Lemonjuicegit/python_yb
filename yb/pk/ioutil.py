from openpyxl import load_workbook
from pandas import read_excel, DataFrame, concat
from pk import decoratorsFunc as dec
from os import path


class IoUtil:

    def __init__(self, gjb_path="挂接表.xlsx", zd_path="zd.xlsx", zrz_path="zrz.xlsx"):
        self.gjb_path = gjb_path
        self.zd_path = zd_path
        self.zrz_path = zrz_path
        self.results = None  # 传递结果信息

        """path文件路径，i修改第几行，string需要修改的内容参数,该方法按行修改内容,并将修改后的内容返回"""

    def revise_file(salf, string, i, file_path):
        b = 0
        try:
            new_file = open(file_path, 'r')
            file_str = list(new_file)  # 读取的文件字符串列表
            new_file.close()
        except FileNotFoundError:
            new_file = open(file_path, 'w')
        try:
            if isinstance(string, str):
                file_str[i] = string + "\n"
            else:
                for a in i:
                    file_str[a] = string[b] + "\n"
                    b += 1
            for st in file_str:
                new_file.write(st)
            new_file.close()
        except IndexError:
            print("请检查是否有改行内容")
        except OSError:
            print("请检查文件路径是否正确")

        return file_str

    # 创建一个新的文件，path:需要创建文件的路径，string（可选）:创建文件时需要写入的文本内容
    def new_file_str(self, path, string=None):
        try:
            if not string:
                op = open(path, "w")
            elif isinstance(string, list):
                op = open(path, "w")
                for i in string:
                    op.write(i)
            elif isinstance(string, str):
                op = open(path, "w")
                op.write(string)
            else:
                print("不是文本数据")

            op.close()
        except FileNotFoundError:
            print("文件路径是否正确")
        except OSError:
            print("请检查文件路径是否正确")

    # 打开excel将一个字段数据转化成列表
    def list_excel(self, field: str):
        self.exc = read_excel(self.gjb_path, dtype="str")
        liste = list(self.exc[field])

        return liste

    def row_excel(self, row, sheet="Sheet1"):
        try:
            exc = read_excel(self.gjb_path, sheet_name=sheet, dtype=str)
            liste = list(exc.loc[row])
        except OSError:
            print("请检查文件路径是否正确")
        return liste

    # x坐标小数位数不够的添零
    def x(self, temp_excel):

        j = 0
        for s in temp_excel:
            if len(str(s)) == 7:
                temp_excel[j] = str(s) + ".000000"
            else:
                temp_excel[j] = str(s) + "0" * (14 - len(str(s)))
            j += 1
        return temp_excel

    # y坐标小数位不够的添零
    def y(self, temp_excel):

        j = 0
        for s in temp_excel:
            if len(str(s)) == 8:
                temp_excel[j] = str(s) + ".000000"
            else:
                temp_excel[j] = str(s) + "0" * (15 - len(str(s)))
            j += 1
        return temp_excel

    # 提取x,y坐标值
    def x_y(self, path, sheet="Sheet1"):

        exc = read_excel(path, sheet_name=sheet)
        bool_excel = list(exc["Unnamed: 2"])
        excel_list = DataFrame()
        i = 8
        while str(bool_excel[i]) != "nan":
            exc1 = exc.iloc[[i], [2, 3]]
            excel_list = concat([excel_list, exc1])
            i += 2
        return excel_list

    # 格式化宗地坐标exf
    def exf_x_y(self, temp_excel, temp2_excel):

        coordinate = []
        j = 0
        while j < len(temp_excel):
            coordinate.append(str(temp2_excel[j]) + "∴" + str(temp_excel[j]) + "∴0.000000\n")
            j += 1
        return coordinate

    # string：excel表列名（A,B,C....）,sheet:工作表名称
    def excel_field(self, string, sheet_name="Sheet1"):

        field = []
        i = 2
        execl_gjb = load_workbook(self.gjb_path, data_only=True)
        sheets = execl_gjb[sheet_name]
        while sheets[string + str(i)].value != None:
            field.append(sheets[string + str(i)].value)
            i += 1
        return field

    # 汇总表
    @dec.getexceptionreturn
    def hzb(self, save_path):
        try:
            data = read_excel(self.gjb_path, dtype="str")
        except FileNotFoundError:
            print("挂接表打不开！")
        except PermissionError:
            print("新增统计表是不是已经打开了！")
        iof = data.loc[:,
              ["乡镇（街道）", "村名", "共用宗权利人",
               "姓名", "身份证", "坐落", "宗地代码",
               "不动产单元号", "旧地籍号", "宗地面积",
               "建筑总面积", "层数", "结构", "登记方式"]]
        iof = iof.set_index("乡镇（街道）")
        iof = iof.rename(columns={"姓名": "办证权利人", "身份证": "证件号码"})

        iof.to_excel(path.join(save_path, "新增统计表.xlsx"))

        results = "汇总表生成成功！"
        return results

    @dec.getexceptionreturn
    def exf(self, save_path):
        op = open(r"templet\exftemplet.exf")
        list_exf = list(op)
        op.close()

        # excel中提取的坐标数据
        list_zddm = self.list_excel("宗地代码")
        list_bdcdyh = self.list_excel("旧地籍号")
        list_xm = self.list_excel("姓名")
        list_zl = self.list_excel("坐落")
        list_zdmj = self.list_excel("宗地面积")
        list_cjr = self.list_excel("调查人")
        # 生成exf
        while_x = 0  # 循环总数
        while while_x < len(list_zddm):
            t = self.x_y("zd.xlsx", list_zddm[while_x])  # 获取宗地的坐标数据
            t2 = self.x_y("zrz.xlsx", list_zddm[while_x])  # 获取幢的坐标数据

            # 填充坐标后面的0
            temp_excel = list(t["Unnamed: 2"])
            temp_excel = self.x(temp_excel)
            temp2_excel = list(t["Unnamed: 3"])
            temp2_excel = self.y(temp2_excel)
            z_excel = list(t2["Unnamed: 2"])
            z_excel = self.x(z_excel)
            z2_excel = list(t2["Unnamed: 3"])
            z2_excel = self.y(z2_excel)

            z_d = self.exf_x_y(temp_excel, temp2_excel)  # 格式化宗地坐标
            z_z = self.exf_x_y(z_excel, z2_excel)  # 格式化自然幢坐标
            try:
                exf1 = "0.0∴0.0∴112500000∴地表宗地∴%s∴%s∴0.0∴0.0∴%s∴%s∴%s∴∴∴∴0∴%s∴∴宅基地∴%s∴∴∴0∴0∴0∴∴1∴112512000\n" % (
                list_bdcdyh[while_x], list_zddm[while_x], list_zl[while_x], list_zdmj[while_x], list_cjr[while_x],
                list_xm[
                    while_x], list_zl[while_x])
                exf2 = "0.0∴0.0∴∴∴∴∴0.0∴%s∴%sF0001∴0.0∴0.0∴∴∴∴∴∴∴∴∴∴∴0∴∴∴2∴2110\n" % (list_bdcdyh[while_x], list_zddm[
                    while_x])
            except TypeError:
                break

            list_exf[1054] = str(len(z_d)) + "\n"
            list_exf[1068] = str(len(z_z)) + "\n"
            list_exf[1109] = exf1
            list_exf[1122] = exf2

            # 存放坐标数据
            n = 0
            zd = 1055
            while n < len(z_d):
                list_exf.insert(zd, z_d[n])
                zd += 1
                n += 1

            n2 = 0
            z = int(zd + 14)
            while n2 < len(z_z):
                list_exf.insert(z, z_z[n2])
                z += 1
                n2 += 1
            self.new_file_str(r"%s\%s%s\%s%s上传.exf" % (
                save_path, list_zddm[while_x], list_xm[while_x], list_zddm[while_x], list_xm[while_x]), list_exf)

            # 删除坐标数据还原模板文件
            l = 1
            while l <= len(z_d):
                di = zd - l
                del list_exf[di]
                l += 1

            l2 = 0
            while l2 < len(z_z):
                z_l2 = int(1068 + len(z_z) - l2)
                del list_exf[z_l2]
                l2 += 1

            while_x += 1

        results = "exf文件生成：%s个 " % (while_x)
        return results

    @dec.getexceptionreturn
    def jzb(self, save_path):
        zddm = self.excel_field("A", "Sheet1")
        zdmj = self.excel_field("P", "Sheet1")
        zjmj = self.excel_field("R", "Sheet1")
        xm = self.excel_field("c", "Sheet1")
        try:
            zd = load_workbook(r".\zd.xlsx", data_only=True, read_only=False)
        except:
            results = "宗地界址表打不开！！！！"
            return results

        n = 0
        for e in zddm:
            try:
                zd[e].cell(row=5, column=1, value="    宗地面积(平方米):%s              坐标系: 2000国家大地坐标系" % zdmj[n])
                zd[e].cell(row=6, column=1, value="    建筑面积(平方米):%s " % zjmj[n])
            except KeyError:
                print(e + "没找到这个表")
                continue

            zd.save(path.join(r"%s\%s%s" % (save_path, e, xm[n]), e + "界址点成果表.xlsx"))
            zd2 = load_workbook(path.join(r"%s\%s%s" % (save_path, e, xm[n]), e + "界址点成果表.xlsx"), data_only=True)
            # 将复制出来的表多余的删除
            for i in zddm:
                if i != e:
                    zd2.remove(zd2[i])

            zd2.save(path.join(r"%s\%s%s" % (save_path, e, xm[n]), e + "界址点成果表.xlsx"))
            n += 1

        results = "界址点成果表导出：%s个" % n
        return results
