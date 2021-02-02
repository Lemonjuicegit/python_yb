from openpyxl import load_workbook
from pandas import read_excel, DataFrame, concat
from pk import decoratorsFunc as dec
from os import path, walk
import json

# 解析json文件
def read_json(json_path):
    with open(json_path, encoding="utf-8") as jsonfile:
        jsonvalue = json.load(jsonfile)
    return jsonvalue

def write_json(json_path,jsonstr):
    with open(json_path,encoding="utf-8") as file1:
        jsonvalue = json.load(file1)
        jsonvalue.update(jsonstr)
    with open(json_path,'w',encoding="utf-8") as file2:
        json.dump(jsonvalue, file2,indent=4)

# 获取paper_path文件夹下面文件的路径
def getfilepath(paper_path, file_type):
    file_name = walk(paper_path, topdown=False)
    file_path = []
    for paper__path, paper_name, file_name in file_name:
        for i in file_name:
            if path.splitext(i)[-1] == file_type:
                file_path.append(path.join(paper__path, i))
    return file_path

def read_exf(path):
    value2 = {"宗地点数量": [], "自然幢点数量": [], "宗地代码": [], "宗地面积": [], "地籍号": [], "坐落": [], "权利人姓名": []}
    for i in path:
        try:
            exf = open(i)
            exf_sring = list(exf)
            nu1 = exf_sring[1054]
            nu2 = exf_sring[1068 + int(nu1)]
            value1 = exf_sring[1109 + int(nu1) + int(nu2)]
            value1 = value1.split("∴")
            value2["宗地点数量"].append(int(nu1))
            value2["自然幢点数量"].append(int(nu2))
            value2["宗地代码"].append(value1[5])
            value2["地籍号"].append(value1[4])
            value2["坐落"].append(value1[8])
            value2["宗地面积"].append(float(value1[9]))
            value2["权利人姓名"].append(value1[15])
        except ValueError:
            value2["宗地代码"].append(i)
            value2["权利人姓名"].append("")
            value2["地籍号"].append("")
            value2["坐落"].append("")
            value2["宗地面积"].append("")
            value2["宗地点数量"].append("")
            value2["自然幢点数量"].append("")
        finally:
            exf.close()
    datafrom = DataFrame(value2)
    return datafrom


"""path文件路径，i修改第几行，string需要修改的内容参数,该方法按行修改内容,并将修改后的内容返回"""


def revise_file(string, i, file_path):
    b = 0
    try:
        new_file = open(file_path, 'r')
        file_str = list(new_file)  # 读取的文件字符串列表
        new_file.close()
    except FileNotFoundError:
        new_file = open(file_path, 'w')
    try:
        if isinstance(string, str):
            file_str[i] = string + "\n"
        else:
            for a in i:
                file_str[a] = string[b] + "\n"
                b += 1
        for st in file_str:
            new_file.write(st)
        new_file.close()
    except IndexError:
        print("请检查是否有改行内容")
    except OSError:
        print("请检查文件路径是否正确")

    return file_str


# 创建一个新的文件，path:需要创建文件的路径，string（可选）:创建文件时需要写入的文本内容
def new_file_str(path, string=None):
    try:
        if not string:
            op = open(path, "w")
        elif isinstance(string, list):
            op = open(path, "w")
            for i in string:
                op.write(i)
        elif isinstance(string, str):
            op = open(path, "w")
            op.write(string)
        else:
            print("不是文本数据")

        op.close()
    except FileNotFoundError:
        print("文件路径是否正确")
    except OSError:
        print("请检查文件路径是否正确")


# 打开excel将一个字段数据转化成列表
def list_excel(gjb_path, field: str):
    exc = read_excel(gjb_path, dtype="str")
    liste = list(exc[field])

    return liste


def row_excel(row, gjb_path, sheet="Sheet1"):
    try:
        exc = read_excel(gjb_path, sheet_name=sheet, dtype=str)
        liste = list(exc.loc[row])
    except OSError:
        print("请检查文件路径是否正确")
    return liste


def excel_field(gjb_path, string, sheet_name="Sheet1"):
    field = []
    i = 2
    execl_gjb = load_workbook(gjb_path, data_only=True)
    sheets = execl_gjb[sheet_name]
    while sheets[string + str(i)].value is not None:
        field.append(sheets[string + str(i)].value)
        i += 1
    return field


class IoUtil:

    def __init__(self, zd_path="zd.xlsx", zrz_path="zrz.xlsx"):
        self.zd_path = zd_path
        self.zrz_path = zrz_path

    # x坐标小数位数不够的添零
    def x(self, temp_excel):

        j = 0
        for s in temp_excel:
            if len(str(s)) == 7:
                temp_excel[j] = str(s) + ".000000"
            else:
                temp_excel[j] = str(s) + "0" * (14 - len(str(s)))
            j += 1
        return temp_excel

    # y坐标小数位不够的添零
    def y(self, temp_excel):
        j = 0
        for s in temp_excel:
            if len(str(s)) == 8:
                temp_excel[j] = str(s) + ".000000"
            else:
                temp_excel[j] = str(s) + "0" * (15 - len(str(s)))
            j += 1
        return temp_excel

    # 提取x,y坐标值
    def x_y(self, jzb_path, sheet="Sheet1"):
        jzb = read_excel(jzb_path, sheet_name=sheet)
        bool_excel = list(jzb["Unnamed: 2"])
        excel_list = DataFrame()
        i = 8
        while str(bool_excel[i]) != "nan":
            exc1 = jzb.iloc[[i], [2, 3]]
            excel_list = concat([excel_list, exc1])
            i += 2
        return excel_list

    # 格式化宗地坐标exf
    def exf_x_y(self, temp_excel, temp2_excel):

        coordinate = []
        j = 0
        while j < len(temp_excel):
            coordinate.append(str(temp2_excel[j]) + "∴" + str(temp_excel[j]) + "∴0.000000\n")
            j += 1
        return coordinate

    # 汇总表
    @dec.getexceptionreturn
    def hzb(self, excel_gjb, save_path):
        iof = excel_gjb.loc[:,
              ["乡镇（街道）", "村名", "共用宗权利人",
               "姓名", "身份证", "坐落", "宗地代码",
               "不动产单元号", "旧地籍号", "宗地面积",
               "建筑总面积", "层数", "结构", "登记方式"]]
        iof = iof.set_index("乡镇（街道）")
        iof = iof.rename(columns={"姓名": "办证权利人", "身份证": "证件号码"})

        iof.to_excel(path.join(save_path, "新增统计表.xlsx"))

        results = "汇总表生成成功！"
        return results

    @dec.getexceptionreturn
    def exf(self, excel_gjb, save_path):
        op = open(r"templet\exftemplet.exf")
        list_exf = list(op)
        op.close()

        # excel中提取的坐标数据
        list_zddm = list(excel_gjb["宗地代码"])
        list_bdcdyh = list(excel_gjb["旧地籍号"])
        list_xm = list(excel_gjb["姓名"])
        list_zl = list(excel_gjb["坐落"])
        list_zdmj = list(excel_gjb["宗地面积"])
        list_cjr = list(excel_gjb["调查人"])
        # 生成exf
        while_x = 0  # 循环总数
        while while_x < len(list_zddm):
            t = self.x_y("zd.xlsx", list_zddm[while_x])  # 获取宗地的坐标数据
            t2 = self.x_y("zrz.xlsx", list_zddm[while_x])  # 获取幢的坐标数据

            # 填充坐标后面的0
            temp_excel = list(t["Unnamed: 2"])
            temp_excel = self.x(temp_excel)
            temp2_excel = list(t["Unnamed: 3"])
            temp2_excel = self.y(temp2_excel)
            z_excel = list(t2["Unnamed: 2"])
            z_excel = self.x(z_excel)
            z2_excel = list(t2["Unnamed: 3"])
            z2_excel = self.y(z2_excel)

            z_d = self.exf_x_y(temp_excel, temp2_excel)  # 格式化宗地坐标
            z_z = self.exf_x_y(z_excel, z2_excel)  # 格式化自然幢坐标
            try:
                exf1 = "0.0∴0.0∴112500000∴地表宗地∴%s∴%s∴0.0∴0.0∴%s∴%s∴%s∴∴∴∴0∴%s∴∴宅基地∴%s∴∴∴0∴0∴0∴∴1∴112512000\n" % (
                    list_bdcdyh[while_x], list_zddm[while_x], list_zl[while_x], list_zdmj[while_x], list_cjr[while_x],
                    list_xm[
                        while_x], list_zl[while_x])
                exf2 = "0.0∴0.0∴∴∴∴∴0.0∴%s∴%sF0001∴0.0∴0.0∴∴∴∴∴∴∴∴∴∴∴0∴∴∴2∴2110\n" % (list_bdcdyh[while_x], list_zddm[
                    while_x])
            except TypeError:
                break

            list_exf[1054] = str(len(z_d)) + "\n"
            list_exf[1068] = str(len(z_z)) + "\n"
            list_exf[1109] = exf1
            list_exf[1122] = exf2

            # 存放坐标数据
            n = 0
            zd = 1055
            while n < len(z_d):
                list_exf.insert(zd, z_d[n])
                zd += 1
                n += 1

            n2 = 0
            z = int(zd + 14)
            while n2 < len(z_z):
                list_exf.insert(z, z_z[n2])
                z += 1
                n2 += 1
            new_file_str(r"%s\%s%s\%s%s上传.exf" % (
                save_path, list_zddm[while_x], list_xm[while_x], list_zddm[while_x], list_xm[while_x]), list_exf)

            # 删除坐标数据还原模板文件
            l = 1
            while l <= len(z_d):
                di = zd - l
                del list_exf[di]
                l += 1

            l2 = 0
            while l2 < len(z_z):
                z_l2 = int(1068 + len(z_z) - l2)
                del list_exf[z_l2]
                l2 += 1

            while_x += 1

        results = "exf文件生成：%s个 " % (while_x)
        return results

    @dec.getexceptionreturn
    def jzb(self, gjb_path, save_path):
        zddm = excel_field(gjb_path, "A", "Sheet1")
        zdmj = excel_field(gjb_path, "P", "Sheet1")
        zjmj = excel_field(gjb_path, "R", "Sheet1")
        xm = excel_field(gjb_path, "c", "Sheet1")
        try:
            zd = load_workbook(r".\zd.xlsx", data_only=True, read_only=False)
        except:
            results = "宗地界址表打不开！！！！"
            return results
        n = 0
        for e in zddm:
            try:
                zd[e].cell(row=5, column=1, value="    宗地面积(平方米):%s              坐标系: 2000国家大地坐标系" % zdmj[n])
                zd[e].cell(row=6, column=1, value="    建筑面积(平方米):%s " % zjmj[n])
            except KeyError as e:
                results = str(e)
                return results

            zd.save(path.join(r"%s\%s%s" % (save_path, e, xm[n]), e + "界址点成果表.xlsx"))
            zd2 = load_workbook(path.join(r"%s\%s%s" % (save_path, e, xm[n]), e + "界址点成果表.xlsx"), data_only=True)
            # 将复制出来的表多余的删除
            for i in zddm:
                if i != e:
                    zd2.remove(zd2[i])

            zd2.save(path.join(r"%s\%s%s" % (save_path, e, xm[n]), e + "界址点成果表.xlsx"))
            n += 1

        results = "界址点成果表导出：%s个" % n
        return results


if __name__ == '__main__':
    a={'y':5}
    write_json(r"..\data.json",a)
