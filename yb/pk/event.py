import socket

from openpyxl import load_workbook
from pandas import *
from xlrd import *


class IoFile:
    """path文件路径，i修改第几行，string需要修改的内容参数,该方法按行修改内容,并将修改后的内容返回"""

    def revise_file(salf, string, i, path):
        b = 0
        try:
            new_file = open(path, 'r')
            file_str = list(new_file)  # 读取的文件字符串列表
            new_file.close()
        except FileNotFoundError:
            new_file = open(path, 'w')
        try:
            if isinstance(string, str):
                file_str[i] = string + "\n"
            else:
                for a in i:
                    file_str[a] = string[b] + "\n"
                    b += 1
            for st in file_str:
                new_file.write(st)
            new_file.close()
        except IndexError:
            print("请检查是否有改行内容")
        except OSError:
            print("请检查文件路径是否正确")

        return file_str

    # 创建一个新的文件，path:需要创建文件的路径，string（可选）:创建文件时需要写入的文本内容
    def new_file_str(self, path, string=None):
        try:
            if not string:
                op = open(path, "w")
                op.close()
            elif isinstance(string, list):
                op = open(path, "w")
                for i in string:
                    op.write(i)
                op.close()
            elif isinstance(string, str):
                op = open(path, "w")
                op.write(string)
                op.close()
            else:
                print("不是文本数据")
        except FileNotFoundError:
            print("文件路径是否正确")
        except OSError:
            print("请检查文件路径是否正确")

    # 打开excel将一个字段数据转化成列表
    def list_excel(salf, path, field: str):

        try:
            salf.exc = read_excel(path)
            liste = list(salf.exc[field])
            return liste
        except OSError:
            print("请检查文件路径是否正确")

    # x坐标小数位数不够的添零
    def x(self, temp_excel):

        j = 0
        for t in temp_excel:
            if len(str(t)) == 7:
                temp_excel[j] = str(t) + ".000000"
            elif len(str(t)) == 9:
                temp_excel[j] = str(t) + "00000"
            elif len(str(t)) == 10:
                temp_excel[j] = str(t) + "0000"
            elif len(str(t)) == 11:
                temp_excel[j] = str(t) + "000"
            elif len(str(t)) == 12:
                temp_excel[j] = str(t) + "00"
            elif len(str(t)) == 13:
                temp_excel[j] = str(t) + "0"
            j += 1
        return temp_excel

    # y坐标小数位不够的添零
    def y(self, temp_excel):

        j = 0
        for t in temp_excel:
            if len(str(t)) == 8:
                temp_excel[j] = str(t) + ".000000"
            elif len(str(t)) == 10:
                temp_excel[j] = str(t) + "00000"
            elif len(str(t)) == 11:
                temp_excel[j] = str(t) + "0000"
            elif len(str(t)) == 12:
                temp_excel[j] = str(t) + "000"
            elif len(str(t)) == 13:
                temp_excel[j] = str(t) + "00"
            elif len(str(t)) == 14:
                temp_excel[j] = str(t) + "0"
            j += 1
        return temp_excel

    # 提取x,y坐标值
    def x_y(self, path, sheet="Sheet1"):

        exc = read_excel(path, sheet_name=sheet)
        bool_excel = list(exc["Unnamed: 2"])
        self.excel = DataFrame()
        i = 8
        while str(bool_excel[i]) != "nan":
            exc1 = exc.iloc[[i], [2, 3]]
            self.excel = concat([self.excel, exc1])
            i += 2
        return (i - 8) / 2

    # 格式化宗地坐标exf
    def exf_x_y(self, temp_excel, temp2_excel):

        exf_x_y_zd = []
        j = 0
        while j < len(temp_excel):
            exf_x_y_zd.append(str(temp2_excel[j]) + "∴" + str(temp_excel[j]) + "∴0.000000\n")
            j += 1
        return exf_x_y_zd

    def key(self):  # 加密方法
        name = socket.gethostname()
        ip = socket.gethostbyname(name)
        key = 1
        n = 0
        while n < len(ip):
            try:
                key = key * ord(name[n]) + ord(ip[n])
                n += 1
            except IndexError:
                break
        return key

    def excel_field(path, string, sheet_name="Sheet1"):  # path:excel表路径, string：excel表列名（A,B,C....）,sheet:工作表名称

        field = []
        i = 2
        execl_gjb = load_workbook(path, data_only=True)
        sheets = execl_gjb[sheet_name]
        while sheets[string + str(i)].value != None:
            field.append(sheets[string + str(i)].value)
            i += 1
        return field

    def hzb(self):  # 汇总表
        try:
            data = read_excel("挂接表.xlsx", dtype="str")
        except FileNotFoundError:
            print("挂接表打不开！")
        except PermissionError:
            print("新增统计表是不是已经打开了！")
        iof = data.loc[:,
              ["乡镇（街道）", "村名", "共用宗权利人",
               "姓名", "身份证", "坐落", "宗地代码",
               "不动产单元号", "旧地籍号", "宗地面积",
               "建筑总面积", "层数", "结构", "登记方式"]]
        iof = iof.set_index("乡镇（街道）")
        iof = iof.rename(columns={"姓名": "办证权利人", "身份证": "证件号码"})
        iof.to_excel("新增统计表.xlsx")
        print("\n汇总表生成成功！")

    def exf(salf):

        iofile = IoFile()
        iofile2 = IoFile()
        op = open("模板.exf")
        list_exf = list(op)
        op.close()

        # excel中提取的坐标数据
        list_zddm = iofile.list_excel("挂接表.xlsx", "宗地代码")
        list_bdcdyh = iofile.list_excel("挂接表.xlsx", "旧地籍号")
        list_xm = iofile.list_excel("挂接表.xlsx", "姓名")
        list_zl = iofile.list_excel("挂接表.xlsx", "坐落")
        list_zdmj = iofile.list_excel("挂接表.xlsx", "宗地面积")
        list_cjr = iofile.list_excel("挂接表.xlsx", "调查人")

        # 生成exf
        while_x = 0  # 循环总数
        err = 0  # 未生成数
        sc = 0  # 生成数
        while while_x < len(list_zddm):

            rate = "进度: " + str(round((while_x + 1) / len(list_zddm), 4) * 100) + "%"  # rate进度
            print("\r" + rate, end="")
            try:
                t = iofile.x_y("zd.xlsx", list_zddm[while_x])  # 获取宗地的坐标数据
                t2 = iofile2.x_y("zrz.xlsx", list_zddm[while_x])  # 获取幢的坐标数据
            except XLRDError:
                while_x += 1
                err += 1
                continue

            # 填充坐标后面的0
            temp_excel = list(iofile.excel["Unnamed: 2"])
            temp_excel = iofile.x(temp_excel)
            temp2_excel = list(iofile.excel["Unnamed: 3"])
            temp2_excel = iofile.y(temp2_excel)
            z_excel = list(iofile2.excel["Unnamed: 2"])
            z_excel = iofile2.x(z_excel)
            z2_excel = list(iofile2.excel["Unnamed: 3"])
            z2_excel = iofile2.y(z2_excel)

            z_d = iofile.exf_x_y(temp_excel, temp2_excel)  # 格式化宗地坐标exf
            z_z = iofile2.exf_x_y(z_excel, z2_excel)  # 格式化宗地坐标exf
            try:
                exf1 = "0.0∴0.0∴112500000∴地表宗地∴" + list_bdcdyh[while_x] + "∴" + list_zddm[while_x] + "∴0.0∴0.0∴" + \
                       list_zl[while_x] + "∴" + str(list_zdmj[while_x]) + "∴" + list_cjr[while_x] + "∴∴∴∴0∴" + list_xm[
                           while_x] + "∴∴宅基地∴" + list_zl[while_x] + "∴∴∴0∴0∴0∴∴1∴112512000\n"
                exf2 = "0.0∴0.0∴∴∴∴∴0.0∴" + list_bdcdyh[while_x] + "∴" + list_zddm[
                    while_x] + "F0001∴0.0∴0.0∴∴∴∴∴∴∴∴∴∴∴0∴∴∴2∴2110\n"
            except TypeError:
                print("挂接表可能没填好！")
                break

            list_exf[1054] = str(int(t)) + "\n"
            list_exf[1068] = str(int(t2)) + "\n"
            list_exf[1109] = exf1
            list_exf[1122] = exf2

            # 存放坐标数据
            n = 0
            zd = 1055
            while n < len(z_d):
                list_exf.insert(zd, z_d[n])
                zd += 1
                n += 1

            n2 = 0
            z = int(zd + 14)
            while n2 < len(z_z):
                list_exf.insert(z, z_z[n2])
                z += 1
                n2 += 1
            iofile.new_file_str(list_zddm[while_x] + list_xm[while_x] + ("上传.exf"), list_exf)

            # 删除坐标数据还原模板文
            l = 1
            while l <= len(z_d):
                di = zd - l
                del list_exf[di]
                l += 1

            l2 = 0
            while l2 < len(z_z):
                z_l2 = int(1068 + t2 - l2)
                del list_exf[z_l2]
                l2 += 1

            sc += 1
            while_x += 1

        create = "    生成：" + str(sc) + "个"  # create产生
        not_create = "    未生成：" + str(err) + "个"  # not_create未产生
        print(create + not_create)

    def jzb(self):

        zddm = IoFile.excel_field("挂接表.xlsx", "A", "Sheet1")
        zdmj = IoFile.excel_field("挂接表.xlsx", "P", "Sheet1")
        zjmj = IoFile.excel_field("挂接表.xlsx", "R", "Sheet1")
        zd = load_workbook("zd.xlsx", data_only=True, read_only=False)

        n = 0
        for e in zddm:
            try:
                zd[e].cell(row=5, column=1, value="    宗地面积(平方米): " + str(zdmj[n]) + "              坐标系: 2000国家大地坐标系")
                zd[e].cell(row=6, column=1, value="    建筑面积(平方米): " + str(zjmj[n]))
            except KeyError:
                print(e + "没找到这个表")
                continue

            zd.save(e + "界址点成果表.xlsx")
            zd2 = load_workbook(e + "界址点成果表.xlsx", data_only=True)
            for i in zddm:
                if i != e:
                    zd2.remove(zd2[i])
            n += 1
            zd2.save(e + "界址点成果表.xlsx")
        print("导出成功！")